import requests
import json
import datetime
import openpyxl
import os
import time
import logging
logging.basicConfig(filename='app.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s',encoding='utf-8')

file_path = os.path.dirname(os.path.abspath(__file__))

def get_flight_data(airport_code,mode,file_path_with_ip):
    page = 1
    all_flights = []
    timestamp = datetime.datetime.now().timestamp()
    while True:
        logging.info("开始获取航班数据")
        url = f"https://api.flightradar24.com/common/v1/airport.json?code={airport_code}&plugin[]=&plugin-setting[schedule][mode]={mode}&plugin-setting[schedule][timestamp]={timestamp}&page={page}&limit=100&fleet=&token="
        
        # 设置代理
        proxies = {
            "http": "http://127.0.0.1:10809",
            "https": "http://127.0.0.1:10809",
        }

        # 设置User-Agent头
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0"
        }
        logging.info("正在发送请求")
        response = requests.get(url, proxies=proxies, headers=headers)
        logging.info("状态码: %s", response.status_code)
        # print("状态码:", response.status_code)
        fail_count = 0
        if response.status_code == 200:
            logging.info("请求成功，正在写入响应到文件")
            print("请求成功")
            data = response.json()  # 将响应内容转换为JSON格式
            flights = data['result']['response']['airport']['pluginData']['schedule'][f'{mode}']['data']
            if len(flights) < 100:  # 如果获取的航班数量小于100，停止请求
                all_flights.extend(flights)
                logging.info("写入完成，结束获取航班数据")
                break
            else:
                logging.info("请求下一页，继续获取航班数据")
                all_flights.extend(flights)
        else:
            logging.error("请求失败，结束获取航班数据")
            print("请求失败")
            fail_count += 1  # 请求失败，计数器加1
            if fail_count >= 5:  # 如果连续失败5次
                return "Error: Too many consecutive failures", 500  # 返回错误信息
            continue
            
        page += 1  # 增加页数，获取下一页的航班数据

    os.makedirs(f'{file_path_with_ip}', exist_ok=True)
    with open(f'{file_path_with_ip}\\{airport_code}_{mode}_flights.json', 'a') as f:
        f.truncate(0)  # 清空文件内容
        json.dump(all_flights, f, indent=4)  # 将航班数据写入JSON文件
    return page

def extract_departure_aircraft_type(data_path, workbook, filter):
    with open(data_path, 'r') as f:
        all_flights = json.load(f)
        
    interested_aircraft = []
    normal_aircraft = []
    
    for flight in all_flights:
        try:
            model_code = flight['flight']['aircraft']['model']['code']
        except:
            print(flight['flight']['aircraft']['model'])
            continue
        departure_time = flight['flight']['time']['scheduled']['departure']
        # 转化为北京时间
        departure_time = datetime.datetime.fromtimestamp(departure_time).strftime('%Y-%m-%d %H:%M:%S')
        
        if any(substring in model_code for substring in filter):
            interested_aircraft.append(flight)
        else:
            normal_aircraft.append(flight)
    
    sheet = workbook.create_sheet("Departure Flight Data")
    sheet.append(["感兴趣的航班"])
    sheet.append(["航班号", "计划起飞时间", "目的地机场", "注册号", "机型"])
    for flight in interested_aircraft:
        
        flight_number = flight['flight']['identification']['number']['default']
        departure_time = datetime.datetime.fromtimestamp(flight['flight']['time']['scheduled']['departure']).strftime('%Y-%m-%d %H:%M:%S')
        destination_airport = flight['flight']['airport']['destination']['code']['iata']
        registration = flight['flight']['aircraft']['registration']
        code = flight['flight']['aircraft']['model']['code']
        
        sheet.append([flight_number, departure_time, destination_airport, registration if registration else 'None', code if code else None])

    sheet.append([])
    sheet.append(["其他航班"])
    sheet.append(["航班号", "计划起飞时间", "目的地机场", "注册号", "机型"])
    for flight in normal_aircraft:
        
        flight_number = flight['flight']['identification']['number']['default']
        departure_time = datetime.datetime.fromtimestamp(flight['flight']['time']['scheduled']['departure']).strftime('%Y-%m-%d %H:%M:%S')
        destination_airport = flight['flight']['airport']['destination']['code']['iata']
        registration = flight['flight']['aircraft']['registration']
        code = flight['flight']['aircraft']['model']['code']
        
        sheet.append([flight_number, departure_time, destination_airport, registration if registration else 'None', code if code else None])

def extract_arrival_aircraft_type(data_path, workbook, filter):
    with open(data_path, 'r') as f:
        all_flights = json.load(f)
        
    interested_aircraft = []
    normal_aircraft = []
    
    for flight in all_flights:
        model_code = flight['flight']['aircraft']['model']['code']
        arrivals_time = flight['flight']['time']['scheduled']['arrival']
        # 转化为北京时间
        arrivals_time = datetime.datetime.fromtimestamp(arrivals_time).strftime('%Y-%m-%d %H:%M:%S')
        
        if any( model_code is None or substring in model_code for substring in filter):
            interested_aircraft.append(flight)
        else:
            normal_aircraft.append(flight)
    
    sheet = workbook.create_sheet("Arrival Flight Data")
    sheet.append(["感兴趣的航班"])
    sheet.append(["航班号", "计划到达时间", "起飞地机场", "注册号", "机型"])
    for flight in interested_aircraft:
        
        flight_number = flight['flight']['identification']['number']['default']
        arrival_time = datetime.datetime.fromtimestamp(flight['flight']['time']['scheduled']['arrival']).strftime('%Y-%m-%d %H:%M:%S')
        origin_airport = flight['flight']['airport']['origin']['code']['iata']
        registration = flight['flight']['aircraft']['registration']
        code = flight['flight']['aircraft']['model']['code']
        
        sheet.append([flight_number, arrival_time, origin_airport, registration if registration else 'None', code if code else None])

    sheet.append([])
    sheet.append(["其他航班"])
    sheet.append(["航班号", "计划到达时间", "起飞地机场", "注册号", "机型"])
    for flight in normal_aircraft:
        
        flight_number = flight['flight']['identification']['number']['default']
        arrival_time = datetime.datetime.fromtimestamp(flight['flight']['time']['scheduled']['arrival']).strftime('%Y-%m-%d %H:%M:%S')
        origin_airport = flight['flight']['airport']['origin']['code']['iata']
        registration = flight['flight']['aircraft']['registration']
        code = flight['flight']['aircraft']['model']['code']
        
        sheet.append([flight_number, arrival_time, origin_airport, registration if registration else 'None', code if code else None])


def getjson(airport_code, data_mode,ip_address):#返回json的函数
    file_path_with_ip = f'{file_path}\\datas\\temp\\{ip_address}\\{airport_code}'
    for d in range(len(data_mode)):
        pages = get_flight_data(airport_code, data_mode[d],file_path_with_ip)
        #判断请求是否太多，是则进行休眠
        if not isinstance(pages, int): #如果返回的不是整数（返回的是尝试失败过多）
            return "Error: Too many consecutive failures", 500 #返回错误信息
        elif pages > 15:
            time.sleep(3)
            print('hold on')
        else:
            continue

    return f'{file_path_with_ip}\\{airport_code}_arrivals_flights.json',f'{file_path_with_ip}\\{airport_code}_departures_flights.json'

'''def run(airport_code, data_mode, jet_type, output_filename,ip_address):#返回excel的函数
    file_path_with_ip = f'{file_path}\\datas\\temp\\{ip_address}\\{airport_code}'
    for d in range(len(data_mode)):
        pages = get_flight_data(airport_code, data_mode[d],file_path_with_ip)
        #判断请求是否太多，是则进行休眠
        if not isinstance(pages, int): #如果返回的不是整数（返回的是尝试失败过多）
            return "Error: Too many consecutive failures", 500 #返回错误信息
        elif pages > 15:
            time.sleep(3)
            print('hold on')
        else:
            continue
    # 写入数据到excel
    workbook = openpyxl.Workbook()
    extract_departure_aircraft_type(f'{file_path_with_ip}\\{airport_code}_departures_flights.json', workbook, jet_type)
    extract_arrival_aircraft_type(f'{file_path_with_ip}\\{airport_code}_arrivals_flights.json', workbook, jet_type)
    del workbook[workbook.sheetnames[0]]
    final_path = f'{file_path_with_ip}\\{output_filename}'
    workbook.save(f'{final_path}')
    return f'{file_path_with_ip}\\{airport_code}_departures_flights.json'''
    
# #测试数据
# airport_code = "xmn"
# data_mode = ['departures','arrivals']
# output_filename = f'{airport_code}.xlsx'
# jet_type = ["A33",'A34',"A35","A38","B78","B77","B76","B74","B75","B78","C9",'74','76','75','77','78','33','35','380','C27']

# #RUN
# run(airport_code, data_mode, jet_type, output_filename)

# 路径
# file_path = os.path.abspath('fr24爬虫test.py')